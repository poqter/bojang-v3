import streamlit as st
import openpyxl
from io import BytesIO
from datetime import datetime

# ✅ 페이지 설정
st.set_page_config(page_title="보장 분석 도우미", layout="centered")

# ✅ 비밀번호 인증 (secrets.toml에서 불러오기)
PASSWORD = st.secrets["PASSWORD"]

st.title("🔐 보장 분석 도우미")
password_input = st.text_input("비밀번호를 입력하세요", type="password")

if password_input != PASSWORD:
    st.warning("비밀번호가 올바르지 않습니다. 올바른 비밀번호를 입력하세요.")
    st.stop()

st.success("✅ 인증 성공! 계속 진행하세요.")

st.write("`컨설팅보장분석.xlsx` 파일을 업로드하면 자동으로 결과 분석 파일이 생성됩니다.")

# ✅ 컨설팅보장분석.xlsx 업로드
uploaded_main = st.file_uploader("⬆️ 컨설팅보장분석.xlsx 파일을 업로드하세요", type=["xlsx"])

# ✅ 개인용보장분석폼폼.xlsx (선택 업로드)
uploaded_print = st.file_uploader("🖨️ (선택) 개인용보장분석폼폼.xlsx 파일을 업로드하세요", type=["xlsx"])

# ✅ print.xlsx 로드
try:
    if uploaded_print:
        print_wb = openpyxl.load_workbook(uploaded_print)
        st.info("✅ 업로드한 print.xlsx를 사용합니다.")
    else:
        print_wb = openpyxl.load_workbook("print.xlsx")
        st.info("📌 기본 내장된 print.xlsx를 사용합니다.")
    print_ws = print_wb.active
except Exception as e:
    st.error(f"❌ print.xlsx 파일을 열 수 없습니다: {e}")
    st.stop()

# ✅ main.xlsx 처리
if uploaded_main:
    try:
        main_wb = openpyxl.load_workbook(uploaded_main, data_only=True)
        main_ws1 = main_wb["계약사항"]
        main_ws2 = main_wb["보장사항"]

        for idx in range(27):
            print_ws.cell(row=10, column=4 + idx).value = main_ws1[f"J{9+idx}"].value

        for row_offset, col in enumerate(['K', 'L']):
            for idx in range(27):
                print_ws.cell(row=8 + row_offset, column=4 + idx).value = main_ws1[f"{col}{9+idx}"].value

        for row in range(2, 8):
            for col in range(6, 30):
                print_ws.cell(row=row, column=col - 2).value = main_ws2.cell(row=row, column=col).value

        for row in range(9, 46):
            for col in range(6, 30):
                print_ws.cell(row=row + 3, column=col - 2).value = main_ws2.cell(row=row, column=col).value

        name_prefix = (main_ws1["B2"].value or "고객")[:3]
        detail_text = main_ws1["D2"].value or ""
        print_ws["A1"] = f"{name_prefix}님의 기존 보험 보장 분석 {detail_text}"

        today_str = datetime.today().strftime("%Y%m%d")
        filename = f"{name_prefix}님의_보장분석_{today_str}.xlsx"
        output_excel = BytesIO()
        print_wb.save(output_excel)
        output_excel.seek(0)

        st.success("✅ 분석이 완료되었습니다.")
        st.download_button(
            label="📥 결과 파일 다운로드",
            data=output_excel,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"⚠️ 오류가 발생했습니다: {str(e)}")
